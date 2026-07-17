"""
SEBI / AMFI monthly net-flow ingestion -> fact_flow_monthly.

AMC-level net inflow/outflow (this module's original target — SEBI's monthly bulletin,
AMFI's AAUM module) is still PDF/POST-only with no clean endpoint (verified 2026-06-21:
AMFI's AAUM module is POST-only and returns Excel, SEBI returns PDFs) — load_csv()/
load_excel() below ingest that shape if you ever get a normalised export of it by hand:

    amc_name,asset_class,category,month,inflow_cr,outflow_cr,aum_cr

But AMFI's separate Monthly Report (MCR) *is* a real, clean, predictably-URLed Excel
export, verified 2026-07-17 — it just reports industry-wide totals per fund CATEGORY,
not per AMC (see load_amfi_mcr_excel() below, the one actually wired into scheduled
ingestion via scripts/ingest_flows.py). `data/flows_seed.csv` / seed_flows_history.py's
old sample AMC-level data is retired now that real category-level data is live.

    python -m ingestion.sebi_flows data/flows_seed.csv
"""

from __future__ import annotations

import csv
import re
import sys

from .db import connect


def derive_class(category: str) -> str:
    c = (category or "").lower()
    if "equity" in c or "elss" in c:
        return "Equity"
    if "debt" in c or "income" in c or "liquid" in c or "gilt" in c or "money market" in c:
        return "Debt"
    if "hybrid" in c or "balanced" in c:
        return "Hybrid"
    if "solution" in c or "retirement" in c:
        return "Solution"
    return "Other"


def _find_col(headers: list[str], *needles: str):
    for i, h in enumerate(headers):
        if any(n in h for n in needles):
            return i
    return None


def load_excel(path: str, month: str, sheet: str | None = None) -> list[tuple]:
    """
    Parse the AMFI/SEBI monthly flow Excel into the same row shape as the CSV path.

    Column detection is by header substring so it survives minor layout changes.
    `month` is the reporting month (the workbook usually encodes it in the filename
    or a title cell, so we pass it explicitly). Recognised headers include:
    AMC / "Mutual Fund Name", "Scheme Category", "Net Inflow/Outflow", AUM.
    """
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    rows = list(ws.iter_rows(values_only=True))

    headers = None
    hdr_idx = None
    for i, r in enumerate(rows[:25]):
        cells = [str(c).strip().lower() if c is not None else "" for c in r]
        if _find_col(cells, "amc", "mutual fund") is not None and _find_col(cells, "category", "scheme") is not None:
            headers, hdr_idx = cells, i
            break
    if headers is None:
        raise ValueError(f"Could not locate a header row in {path}")

    c_amc = _find_col(headers, "amc", "mutual fund name", "name of the")
    c_cat = _find_col(headers, "category", "scheme type")
    c_in = _find_col(headers, "inflow", "gross inflow", "sales")
    c_out = _find_col(headers, "outflow", "redemption", "repurchase")
    c_net = _find_col(headers, "net inflow", "net flow", "net")
    c_aum = _find_col(headers, "aum", "assets under management", "net assets")

    def num(r, i):
        if i is None or r[i] in (None, "", "-"):
            return None
        try:
            return float(str(r[i]).replace(",", ""))
        except ValueError:
            return None

    out: list[tuple] = []
    for r in rows[hdr_idx + 1:]:
        if c_amc is None or not r[c_amc] or not str(r[c_amc]).strip():
            continue
        amc = str(r[c_amc]).strip()
        category = str(r[c_cat]).strip() if c_cat is not None and r[c_cat] else ""
        inflow, outflow, net, aum = num(r, c_in), num(r, c_out), num(r, c_net), num(r, c_aum)
        if net is None and inflow is not None and outflow is not None:
            net = inflow - outflow
        if net is None:
            continue
        out.append((amc, derive_class(category), category, month,
                    inflow or 0, outflow or 0, round(net, 2), aum))
    return out


# AMFI's monthly MCR (Mutual Fund Categorywise Report) — a real, live, monthly-updated Excel
# export at a predictable URL: https://portal.amfiindia.com/spages/am<mon><year>repo.xls
# (verified 2026-07-17: https://www.amfiindia.com/research-information/amfi-monthly lists every
# month back to 1999 with this exact link pattern). Old-format binary .xls (needs xlrd, not
# openpyxl). Its single sheet ("MCR Monthly Report") is a hierarchical, industry-wide breakdown
# by fund category — NOT broken out per-AMC (checked every row of a real June 2026 download; no
# AMC dimension exists anywhere in the file). So this ingests real per-CATEGORY flow, not
# per-AMC flow — a coarser grain than fact_flow_monthly's schema was designed around, but
# genuinely real and automatable, unlike the AMC-level data load_excel() above expects (which
# has never had a working source — see the module docstring).
#
# Row shape (0-indexed columns), confirmed against the real file: col0=Sr (roman numeral or
# letter), col1=Scheme Name, col4=Funds Mobilized (inflow, INR crore), col5=Repurchase/
# Redemption (outflow, INR crore), col6=Net Inflow/Outflow (INR crore), col7=Net AUM (INR crore).
# Only "A — Open ended Schemes" is ingested: Close-ended and Interval schemes don't have ongoing
# subscriptions/redemptions the way open-ended funds do (no meaningful "monthly flow" concept),
# and are economically negligible next to open-ended (~0.02% of total AUM combined in the real
# file checked). Within Group A, a leaf category row has a lowercase-roman col0 (i, ii, iii, ...);
# uppercase-roman rows (I-V) are group headers (blank value cells) whose Scheme Name doubles as
# the broad asset-class bucket for every leaf beneath it, and "Sub Total"/blank/"Total A..." rows
# are rollups, not data.
_GROUP_TO_ASSET_CLASS = {
    "I": "Debt",
    "II": "Equity",
    "III": "Hybrid",
    "IV": "Solution",
    "V": "Other",
}

AMFI_MCR_SOURCE = "AMFI Monthly Report (MCR)"


def load_amfi_mcr_excel(path: str, month: str) -> list[dict]:
    """Parse the real AMFI MCR .xls into fact_flow_monthly-shaped rows (dicts, ready for a
    Supabase REST upsert). `month` is the reporting month as YYYY-MM-01 (the workbook only
    states it in a title cell, e.g. "Monthly Report for the month of June 2026 ", not a
    structured date column, so it's supplied by the caller from the source URL/filename)."""
    import xlrd

    wb = xlrd.open_workbook(path)
    sh = wb.sheet_by_index(0)

    rows: list[dict] = []
    group = None  # current broad bucket (Debt/Equity/Hybrid/Solution/Other), None once out of Group A
    in_open_ended = False
    for r in range(sh.nrows):
        c0 = str(sh.cell_value(r, 0)).strip()
        c1 = str(sh.cell_value(r, 1)).strip()

        if c0 == "A":
            in_open_ended = True
            continue
        if c0 in ("B", "C"):  # Close Ended / Interval Schemes — out of scope, see module note
            break
        if not in_open_ended:
            continue
        if c1.startswith("Total A"):
            break
        if c0 in _GROUP_TO_ASSET_CLASS:
            group = _GROUP_TO_ASSET_CLASS[c0]
            continue
        if not c0 or not c1 or c1.startswith("Sub Total") or group is None:
            continue
        if not re.fullmatch(r"[ivxlcdm]+", c0, re.IGNORECASE):
            continue  # not a lowercase-roman leaf row (defensive; every real leaf matched this)

        def num(col):
            v = sh.cell_value(r, col)
            return float(v) if isinstance(v, (int, float)) else None

        inflow, outflow, net, aum = num(4), num(5), num(6), num(7)
        if net is None:
            continue
        rows.append(
            {
                "amc_name": "Industry (All AMCs)",
                "asset_class": c1,
                "category": group,
                "month": month,
                "inflow_cr": round(inflow, 2) if inflow is not None else 0,
                "outflow_cr": round(outflow, 2) if outflow is not None else 0,
                "net_flow_cr": round(net, 2),
                "aum_cr": round(aum, 2) if aum is not None else None,
                "source": AMFI_MCR_SOURCE,
            }
        )
    return rows


def load_csv(path: str) -> list[tuple]:
    rows: list[tuple] = []
    with open(path, newline="") as fh:
        for r in csv.DictReader(fh):
            inflow = float(r["inflow_cr"])
            outflow = float(r["outflow_cr"])
            rows.append(
                (
                    r["amc_name"].strip(),
                    r["asset_class"].strip(),
                    (r.get("category") or "").strip() or None,
                    r["month"].strip(),
                    inflow,
                    outflow,
                    round(inflow - outflow, 2),
                    float(r["aum_cr"]) if r.get("aum_cr") else None,
                )
            )
    return rows


def run(path: str) -> int:
    rows = load_csv(path)
    if not rows:
        raise RuntimeError(f"No rows parsed from {path}")
    with connect() as conn:
        with conn.cursor() as cur:
            cur.executemany(
                """
                INSERT INTO fact_flow_monthly
                    (amc_name, asset_class, category, month, inflow_cr, outflow_cr, net_flow_cr, aum_cr)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (amc_name, asset_class, month) DO UPDATE SET
                    inflow_cr   = EXCLUDED.inflow_cr,
                    outflow_cr  = EXCLUDED.outflow_cr,
                    net_flow_cr = EXCLUDED.net_flow_cr,
                    aum_cr      = EXCLUDED.aum_cr
                """,
                rows,
            )
    print(f"Loaded {len(rows)} monthly-flow rows from {path}")
    return len(rows)


if __name__ == "__main__":
    run(sys.argv[1] if len(sys.argv) > 1 else "data/flows_seed.csv")
